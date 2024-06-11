import time
from configparser import ConfigParser

import pandas as pd
import os

from AnalyzeData import AnalyzeData
from C import C

import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formula.translate import Translator
import pysnowball as ball
import subprocess;


def prepare():
    config = ConfigParser()
    config.read('token.config')
    pro = ts.pro_api(config.get('token', 'tushare'))
    ball.set_token(config.get('token', 'xueqiu'))  # xq_a_token=XXXX

    if not os.path.exists(C.DATA_PATH):
        os.mkdir(C.DATA_PATH)

        # 更新基础数据
        DataUpgrade.updateBK(ball)
        DataUpgrade.updateStocks(pro)


if __name__ == '__main__':
    prepare()

    # 今天
    date = time.strftime("%Y-%m-%d", time.localtime())
    kline_data = AnalyzeData.get_days_kline_all_stocks()

    # print(kline_data)

    # 市值100-1000亿
    filtered_df = kline_data[kline_data['market_capital'] > 100 * 10000 * 10000]
    print(f'市值>100亿 {filtered_df.shape[0]}')
    filtered_df = filtered_df[filtered_df['market_capital'] < 1000 * 10000 * 10000]
    print(f'市值<1000亿 {filtered_df.shape[0]}')

    # 成交额>3亿
    filtered_df = filtered_df[filtered_df['amount'] > 3 * 10000 * 10000]
    print(f'成交额>3亿 {filtered_df.shape[0]}')

    # 今日下影
    filtered_df = filtered_df[filtered_df['low'] / filtered_df['last_close'] - 1 < -0.04]
    filtered_df = filtered_df[-0.05 < filtered_df['low'] / filtered_df['last_close'] - 1]
    print(f'下影 {filtered_df.shape[0]}')
    # 今日上影
    filtered_df = filtered_df[0.02 < filtered_df['high'] / filtered_df['last_close'] - 1]
    filtered_df = filtered_df[filtered_df['high'] / filtered_df['last_close'] - 1 < 0.03]
    print(f'上影 {filtered_df.shape[0]}')

    filtered_df = filtered_df.reset_index()

    wb = openpyxl.load_workbook('module.xlsx')
    # 选择要操作的工作表
    sheet = wb.active  # 或者使用 workbook['Sheet1'] 选择特定的工作表
    print(f' {filtered_df}')
    index = 1
    for table_index, row in filtered_df.iterrows():
        # 获取 DataFrame 中的相关数据

        code = row['code']
        volume_ratio = AnalyzeData.get_stock_detail(ball, code).iloc[0,]['volume_ratio']
        if 0.6 < volume_ratio < 1:  # 量比在0.6到1之间
            name = row['name']
            print(code, name)
            price = row['close']
            indicator = abs(
                (row['last_close'] - row['last_low']) / (row['last_open'] - row['last_close']) - 0.382) + abs(
                (row['high'] - row['close']) / (row['close'] - row['open']) - 0.382)
            # if indicator < 0.3: #
            # print(sheet)
            # 将数据填充到 Excel 文件中的相应列
            name_cell = sheet.cell(row=index + 1, column=1, value=f'{name} {str(code)}')
            name_cell.hyperlink = f'https://xueqiu.com/S/{code}'
            sheet.cell(row=index + 1, column=2, value=round(float(indicator), 2))
            sheet.cell(row=index + 1, column=3, value=price)
            index = index + 1

    # # 保存修改后的 Excel 文件
    wb.save(f'{C.XLS_OUTPUT_PATH}short2_{date}.xlsx')
    wb.close()
    subprocess.run(['say', '运行完毕，恭喜发财'])
