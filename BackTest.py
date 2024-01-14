from openpyxl import load_workbook
from AnalyzeData import AnalyzeData
import pysnowball as ball
from configparser import ConfigParser

if __name__ == '__main__':
    # 读取xls文件
    file_path = "/Users/xirtam/Downloads/第一周.xlsx"
    workbook = load_workbook(file_path)
    # 获取第一个工作表
    sheet = workbook.active

    config = ConfigParser()
    config.read('token.config')

    ball.set_token(config.get('token', 'xueqiu'))  # xq_a_token=XXXX

    result_list = []
    # 遍历每行并打印第一列的后8位字符串
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
        first_column_value = str(row[0])
        if first_column_value == '表格 1' or first_column_value == '名称':
            continue
        last_eight_characters = first_column_value[-8:]  # 获取后8位字符
        result_list.append(last_eight_characters)

    print('开始获取行情数据...')
    prices = AnalyzeData.getWeekPercentBySymbols(result_list)

    result_column_index = None
    for col_index, col in enumerate(sheet.iter_cols(min_row=1, max_row=1, values_only=True), start=1):
        if col[0] == '结果':
            result_column_index = col_index
            break

    for index, value in enumerate(prices, start=1):
        sheet.cell(row=index + 2, column=result_column_index, value=value)
    workbook.save(file_path)
    print('补充行情成功')
