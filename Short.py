import time
from configparser import ConfigParser

import pandas as pd
import os

from AnalyzeData import AnalyzeData
from C import C

import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formula.translate import Translator
import pysnowball as ball


def prepare():
    if not os.path.exists(C.DATA_PATH):
        os.mkdir(C.DATA_PATH)
    config = ConfigParser()
    config.read('token.config')

    ball.set_token(config.get('token', 'xueqiu'))  # xq_a_token=XXXX


def test():
    # 创建一个包含数据的 DataFrame
    data = {'F': [3, 6, 8],
            'G': [2, 5, 9],
            'H': [7, 4, 1]}
    df = pd.DataFrame(data)

    # 创建一个 Excel 文件
    wb = Workbook()
    ws = wb.active

    # 将数据写入工作表
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 在工作表中添加公式
    for row in range(2, df.shape[0] + 2):
        formula = f'=INDEX(F$1:H$1, 1, MATCH(MAX(F{row}:H{row}), F{row}:H{row}, 1))'
        ws[f'I{row}'].value = Translator(formula, origin="A1").translate_formula()

    # 保存 Excel 文件
    wb.save('output_with_formula.xlsx')


if __name__ == '__main__':
    prepare()
    date = time.strftime("%Y-%m-%d", time.localtime())

    last_week_activity = AnalyzeData.get_last_week_activity_index()

    # 筛选数据，上周上涨的
    filtered_df = last_week_activity[last_week_activity['last_week_percent'] > 0]

    # 根据 indicator 列从大到小排序，然后取前50条记录
    sorted_filtered_df = filtered_df.sort_values(by='indicator', ascending=False).reset_index()
    # sorted_filtered_df = filtered_df.sort_values(by='indicator', ascending=False).head(100).reset_index()

    # print(sorted_filtered_df)

    wb = openpyxl.load_workbook('module.xlsx')
    # 选择要操作的工作表
    sheet = wb.active  # 或者使用 workbook['Sheet1'] 选择特定的工作表

    # for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
    #     # 遍历每一列
    #     for cell in row:
    #         # 输出每个单元格的值
    #         print(cell.value, end='\t')
    #     print()

    index = 1
    for table_index, row in sorted_filtered_df.iterrows():
        # 获取 DataFrame 中的相关数据
        code = row['code']
        name = row['name']
        indicator = row['indicator']
        price = row['price']

        # print(sheet)
        # 将数据填充到 Excel 文件中的相应列
        name_cell = sheet.cell(row=index + 1, column=1, value=f'{name} {str(code)}')
        name_cell.hyperlink = f'https://xueqiu.com/S/{code}'

        sheet.cell(row=index + 1, column=2, value=round(float(indicator), 2))
        sheet.cell(row=index + 1, column=3, value=price)

        index = index + 1

    # # 保存修改后的 Excel 文件
    wb.save(f'{C.XLS_OUTPUT_PATH}your_updated_excel_file.xlsx')
    wb.close()
