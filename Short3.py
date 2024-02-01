import time
from configparser import ConfigParser

import pandas as pd
import os

from AnalyzeData import AnalyzeData
from C import C

import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formula.translate import Translator
import pysnowball as ball
import subprocess;


def prepare():
    if not os.path.exists(C.DATA_PATH):
        os.mkdir(C.DATA_PATH)
    config = ConfigParser()
    config.read('token.config')

    ball.set_token(config.get('token', 'xueqiu'))  # xq_a_token=XXXX


if __name__ == '__main__':
    prepare()

    # 今天
    date = time.strftime("%Y-%m-%d", time.localtime())
    kline_data = AnalyzeData.get_weeks_low_vol()

    print(kline_data)

    # 筛选数据，今日红
    # filtered_df = kline_data[(kline_data['close'] / kline_data['open']) > 1.01]
    # print(f'今日红{filtered_df.shape[0]}')

    filtered_df = kline_data.reset_index()

    wb = openpyxl.load_workbook('module.xlsx')
    # 选择要操作的工作表
    sheet = wb.active  # 或者使用 workbook['Sheet1'] 选择特定的工作表

    index = 1
    for table_index, row in filtered_df.iterrows():
        # 获取 DataFrame 中的相关数据
        code = row['code']
        name = row['name']
        indicator = row['indicator']
        name_cell = sheet.cell(row=index + 1, column=1, value=f'{name} {str(code)}')
        name_cell.hyperlink = f'https://xueqiu.com/S/{code}'
        sheet.cell(row=index + 1, column=2, value=round(float(indicator), 2))
        index = index + 1

    # # 保存修改后的 Excel 文件
    wb.save(f'{C.XLS_OUTPUT_PATH}short3_{date}.xlsx')
    wb.close()
    subprocess.run(['say', '运行完毕，恭喜发财'])
